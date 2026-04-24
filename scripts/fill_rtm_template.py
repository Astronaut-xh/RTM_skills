#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
RTM模板填写辅助脚本
用于读取输入文档并填写RTM模板
"""

import os
import sys
import json
from pathlib import Path

def read_docx(file_path):
    """读取docx文件内容"""
    try:
        from docx import Document
        doc = Document(file_path)
        content = []
        for para in doc.paragraphs:
            if para.text.strip():
                content.append(para.text)
        return '\n'.join(content)
    except ImportError:
        print("Error: python-docx not installed. Run: pip install python-docx")
        sys.exit(1)

def read_xlsx(file_path, sheet_name=None):
    """读取xlsx文件内容"""
    try:
        import pandas as pd
        if sheet_name:
            return pd.read_excel(file_path, sheet_name=sheet_name)
        else:
            return pd.read_excel(file_path, sheet_name=None)  # 返回所有sheets
    except ImportError:
        print("Error: pandas not installed. Run: pip install pandas openpyxl")
        sys.exit(1)

def write_xlsx(template_path, output_path, dr_fl_data, fl_tp_data):
    """填写模板并保存"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(template_path)

        # 填写DR-FL sheet
        if 'DR-FL' in wb.sheetnames:
            ws_dr_fl = wb['DR-FL']
            for row_idx, row_data in enumerate(dr_fl_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    ws_dr_fl.cell(row=row_idx, column=col_idx, value=value)

        # 填写FL-TP sheet
        if 'FL-TP' in wb.sheetnames:
            ws_fl_tp = wb['FL-TP']
            for row_idx, row_data in enumerate(fl_tp_data, start=3):
                for col_idx, value in enumerate(row_data, start=1):
                    ws_fl_tp.cell(row=row_idx, column=col_idx, value=value)

        wb.save(output_path)
        print(f"Successfully saved to: {output_path}")
        return True
    except Exception as e:
        print(f"Error writing xlsx: {e}")
        return False

def extract_requirements_from_lrs(content):
    """从LRS文档中提取需求项"""
    import re
    requirements = []

    # 匹配需求编号格式，如 APLC.LRS.FUNC.01
    pattern = r'(APLC\.LRS\.[A-Z]+\.\d+)\s+(.+?)(?=APLC\.LRS\.|$)'
    matches = re.findall(pattern, content, re.DOTALL)

    for match in matches:
        req_id = match[0]
        req_content = match[1].strip()
        requirements.append({
            'id': req_id,
            'content': req_content
        })

    return requirements

def map_to_feature_category(content):
    """根据内容关键词映射到Feature类别"""
    keywords_map = {
        '时钟': ['时钟', '频率', 'clk', 'clock', 'CLK'],
        '复位': ['复位', 'reset', 'rst', 'Reset'],
        '寄存器': ['寄存器', 'CSR', 'register', 'Register'],
        '工作模式': ['模式', 'mode', '工作状态', '工作模式'],
        '数据接口': ['数据接口', '数据通路', 'pdi', 'pdo', '数据'],
        '控制接口': ['控制接口', '控制信号', 'opcode', '控制'],
        '配置接口': ['配置', 'config', 'CTRL', '配置接口'],
        '中断': ['中断', 'IRQ', 'interrupt', '中断处理'],
        '异常': ['异常', '错误', 'error', '异常处理', 'ERR'],
        '低功耗': ['低功耗', '功耗', 'power', 'CG', '低功耗'],
        '性能': ['性能', '频率', '延迟', 'latency', '性能'],
        'DFX': ['DFX', '调试', 'debug', '观测', 'DFX'],
        'Memory map': ['memory map', '地址映射', 'Memory map'],
        '总线接口': ['总线', 'AHB', 'AXI', 'bus', '总线接口']
    }

    for category, keywords in keywords_map.items():
        for keyword in keywords:
            if keyword.lower() in content.lower():
                return category

    return '其他'

# 示例用法
if __name__ == "__main__":
    # 示例数据（DR-FL）
    dr_fl_data = [
        ['DR.APLC.001', '时钟', 'FL_001_001', '示例Feature描述', None],
    ]

    # 示例数据（FL-TP）- checker和Testcase编号为空
    fl_tp_data = [
        ['FL_001_001', '时钟', 'TP_001', '条件：test_mode_i=1，en_i=1\n激励：执行时钟相关操作\n期望：符合规格，参考LRS §5.3', None, None],
    ]

    print("RTM Template Filler Script")
    print("Usage: Import and use functions in your code")
